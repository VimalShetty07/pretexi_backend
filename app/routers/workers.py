import io
import secrets
import string
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.database import get_db
from app.core.security import hash_password
from app.core.config import get_settings
from app.routers.deps import get_current_user, require_staff
from sqlalchemy import func as sa_func
from app.models.models import User, Worker, UserRole, DocumentChecklist, ChecklistStatus
from app.schemas.schemas import WorkerCreate, WorkerUpdate, WorkerOut, WorkerDetailOut
from app.routers.documents import create_checklist_for_worker

settings = get_settings()
DEFAULT_EMPLOYEE_PASSWORD = settings.DEFAULT_EMPLOYEE_PASSWORD


def _create_employee_user(
    db: Session,
    worker: Worker,
    organisation_id: str,
) -> User | None:
    """Auto-create a User account with EMPLOYEE role linked to a Worker.
    Returns the User, or None if the worker has no email.
    """
    if not worker.email:
        return None

    existing = db.query(User).filter(User.email == worker.email).first()
    if existing:
        if not existing.worker_id:
            existing.worker_id = worker.id
            existing.role = UserRole.EMPLOYEE
        return existing

    user = User(
        organisation_id=organisation_id,
        email=worker.email,
        hashed_password=hash_password(DEFAULT_EMPLOYEE_PASSWORD),
        full_name=worker.name,
        role=UserRole.EMPLOYEE,
        worker_id=worker.id,
        phone=worker.phone,
    )
    db.add(user)
    return user

router = APIRouter(prefix="/workers", tags=["workers"])

BULK_COLUMNS = [
    "first_name", "last_name", "job_title", "email", "phone", "nationality",
    "department", "soc_code", "salary", "route", "work_location",
    "start_date", "visa_expiry", "passport_expiry", "brp_expiry", "stage",
    "address", "postal_code", "date_of_birth", "place_of_birth", "country_of_birth",
    "gender", "ethnicity", "religion", "ni_number",
    "passport_number", "passport_place_of_issue", "passport_issue_date",
    "emergency_contact_name", "emergency_contact_phone",
    "next_of_kin_name", "next_of_kin_phone",
    "employee_id", "employee_type",
]

BULK_HEADERS = [
    "First Name *", "Last Name *", "Job Title *", "Email", "Phone", "Nationality",
    "Department", "SOC Code", "Salary (£)", "Visa Route", "Work Location",
    "Start Date (YYYY-MM-DD)", "Visa Expiry (YYYY-MM-DD)",
    "Passport Expiry (YYYY-MM-DD)", "BRP Expiry (YYYY-MM-DD)", "Stage",
    "Home Address", "Postal Code", "Date of Birth (YYYY-MM-DD)",
    "Place of Birth", "Country of Birth",
    "Gender", "Ethnicity", "Religion", "NI Number",
    "Passport Number", "Passport Place of Issue", "Passport Issue Date (YYYY-MM-DD)",
    "Emergency Contact Name", "Emergency Contact Phone",
    "Next of Kin Name", "Next of Kin Phone",
    "Employee ID", "Employee Type",
]

SAMPLE_ROWS = [
    ["Claudia Manuel", "Canelhas Pinhao", "Live in Carer", "cpinhao@gmail.com", "7748519695",
     "Portugese", "Care", "6145", 25000, "Skilled Worker", "London",
     "2024-06-01", "2028-05-31", "2030-12-19", "2028-05-31", "active_sponsorship",
     "Rua dos Eucaliptos 98, Bairro da Encarnacao, LISBOA, Portugal", "1800-202",
     "1971-07-18", "Lisboa", "Portugal",
     "Female", "", "", "NJ645681B",
     "CH033822", "Portugal", "2025-12-19",
     "Maria Emilia Salves Canelhas Pinhao", "351936991149",
     "Maria Emilia Salves Canelhas Pinhao", "351936991149",
     "1", "migrant"],
    ["Tom", "Brown", "Data Scientist", "tom.brown@company.com", "+44 7700 900002",
     "Nigerian", "Analytics", "2425", 42000, "Skilled Worker", "Manchester Office",
     "2025-01-15", "2029-01-14", "2031-07-20", "2029-01-14", "recruitment",
     "12 Oxford Road, Manchester", "M1 5QA",
     "1990-03-25", "Lagos", "Nigeria",
     "Male", "", "", "CD789012E",
     "NG654321", "Lagos", "2023-03-25",
     "Funke Brown", "+44 7700 900003",
     "Funke Brown", "+44 7700 900003",
     "2", "migrant"],
]


# ── List workers ──

@router.get("", response_model=list[WorkerOut])
@router.get("/", response_model=list[WorkerOut], include_in_schema=False)
def list_workers(
    stage: str | None = None,
    status: str | None = None,
    search: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.EMPLOYEE:
        if current_user.worker_id:
            worker = db.query(Worker).filter(Worker.id == current_user.worker_id).first()
            return [worker] if worker else []
        return []

    query = db.query(Worker).filter(Worker.organisation_id == current_user.organisation_id)

    if stage:
        query = query.filter(Worker.stage == stage)
    if status:
        query = query.filter(Worker.status == status)
    if search:
        query = query.filter(Worker.name.ilike(f"%{search}%"))

    return query.order_by(Worker.created_at.desc()).all()


# ── Compliance summary for all workers ──

@router.get("/compliance-summary")
def get_compliance_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    rows = (
        db.query(
            DocumentChecklist.worker_id,
            DocumentChecklist.status,
            sa_func.count(DocumentChecklist.id),
        )
        .group_by(DocumentChecklist.worker_id, DocumentChecklist.status)
        .all()
    )

    summary: dict[str, dict] = {}
    for worker_id, st, cnt in rows:
        if worker_id not in summary:
            summary[worker_id] = {"total": 0, "verified": 0, "uploaded": 0, "rejected": 0}
        summary[worker_id]["total"] += cnt
        if st in (ChecklistStatus.VERIFIED, ChecklistStatus.NOT_APPLICABLE):
            summary[worker_id]["verified"] += cnt
        elif st == ChecklistStatus.UPLOADED:
            summary[worker_id]["uploaded"] += cnt
        elif st == ChecklistStatus.REJECTED:
            summary[worker_id]["rejected"] += cnt

    return summary


# ── Template download (must be before /{worker_id}) ──

@router.get("/template", response_class=StreamingResponse)
def download_template(current_user: User = Depends(require_staff)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B5DA8", end_color="2B5DA8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )

    for col_idx, header in enumerate(BULK_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(SAMPLE_ROWS, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for col_idx in range(1, len(BULK_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_bulk_template.xlsx"},
    )


# ── Bulk upload (must be before /{worker_id}) ──

@router.post("/bulk")
def bulk_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")

    try:
        wb = load_workbook(file.file, read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the Excel file. Make sure it is a valid .xlsx")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    if not rows:
        raise HTTPException(status_code=400, detail="The spreadsheet has no data rows")

    created = 0
    errors = []

    for row_num, row in enumerate(rows, 2):
        if not row or not row[0]:
            continue

        values = list(row) + [None] * (len(BULK_COLUMNS) - len(row))

        col = {k: values[i] for i, k in enumerate(BULK_COLUMNS)}

        first_name = str(col["first_name"]).strip() if col["first_name"] else ""
        last_name = str(col["last_name"]).strip() if col["last_name"] else ""
        job_title = str(col["job_title"]).strip() if col["job_title"] else ""

        if (not first_name and not last_name) or not job_title:
            errors.append(f"Row {row_num}: first_name/last_name and job_title are required")
            continue

        full_name = f"{first_name} {last_name}".strip()

        def parse_date(val):
            if val is None:
                return None
            if isinstance(val, datetime):
                return val.replace(tzinfo=timezone.utc) if val.tzinfo is None else val
            try:
                return datetime.strptime(str(val).strip(), "%Y-%m-%d").replace(tzinfo=timezone.utc)
            except ValueError:
                return None

        def opt_str(val):
            return str(val).strip() if val else None

        salary_raw = col["salary"]
        try:
            salary = float(salary_raw) if salary_raw else 0
        except (ValueError, TypeError):
            salary = 0

        worker = Worker(
            organisation_id=current_user.organisation_id,
            name=full_name,
            first_name=first_name or None,
            last_name=last_name or None,
            job_title=job_title,
            email=opt_str(col["email"]),
            phone=opt_str(col["phone"]),
            nationality=opt_str(col["nationality"]),
            department=opt_str(col["department"]),
            soc_code=opt_str(col["soc_code"]),
            salary=salary,
            route=opt_str(col["route"]) or "Skilled Worker",
            work_location=opt_str(col["work_location"]),
            start_date=parse_date(col["start_date"]),
            visa_expiry=parse_date(col["visa_expiry"]),
            passport_expiry=parse_date(col["passport_expiry"]),
            brp_expiry=parse_date(col["brp_expiry"]),
            stage=opt_str(col["stage"]) or "recruitment",
            address=opt_str(col["address"]),
            postal_code=opt_str(col["postal_code"]),
            date_of_birth=parse_date(col["date_of_birth"]),
            place_of_birth=opt_str(col["place_of_birth"]),
            country_of_birth=opt_str(col["country_of_birth"]),
            gender=opt_str(col["gender"]),
            ethnicity=opt_str(col["ethnicity"]),
            religion=opt_str(col["religion"]),
            ni_number=opt_str(col["ni_number"]),
            passport_number=opt_str(col["passport_number"]),
            passport_place_of_issue=opt_str(col["passport_place_of_issue"]),
            passport_issue_date=parse_date(col["passport_issue_date"]),
            emergency_contact_name=opt_str(col["emergency_contact_name"]),
            emergency_contact_phone=opt_str(col["emergency_contact_phone"]),
            next_of_kin_name=opt_str(col["next_of_kin_name"]),
            next_of_kin_phone=opt_str(col["next_of_kin_phone"]),
            employee_id=opt_str(col["employee_id"]),
            employee_type=opt_str(col["employee_type"]),
        )
        db.add(worker)
        db.flush()
        create_checklist_for_worker(db, worker.id)
        _create_employee_user(db, worker, current_user.organisation_id)
        created += 1

    if created > 0:
        db.commit()

    return {
        "created": created,
        "errors": errors,
        "default_password": DEFAULT_EMPLOYEE_PASSWORD if created > 0 else None,
    }


# ── Single worker CRUD (dynamic {worker_id} routes last) ──

@router.get("/{worker_id}", response_model=WorkerDetailOut)
def get_worker(
    worker_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.EMPLOYEE and current_user.worker_id != worker_id:
        raise HTTPException(status_code=403, detail="You can only view your own record")

    worker = db.query(Worker).filter(
        Worker.id == worker_id,
        Worker.organisation_id == current_user.organisation_id,
    ).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    return worker


@router.post("", response_model=WorkerOut, status_code=status.HTTP_201_CREATED)
@router.post("/", response_model=WorkerOut, status_code=status.HTTP_201_CREATED, include_in_schema=False)
def create_worker(
    payload: WorkerCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    data = payload.model_dump()
    if not data.get("name") and (data.get("first_name") or data.get("last_name")):
        data["name"] = f'{data.get("first_name", "") or ""} {data.get("last_name", "") or ""}'.strip()
    worker = Worker(organisation_id=current_user.organisation_id, **data)
    db.add(worker)
    db.flush()
    create_checklist_for_worker(db, worker.id)
    emp_user = _create_employee_user(db, worker, current_user.organisation_id)
    db.commit()
    db.refresh(worker)

    result = WorkerOut.model_validate(worker).model_dump()
    if emp_user:
        result["employee_login"] = {
            "email": emp_user.email,
            "default_password": DEFAULT_EMPLOYEE_PASSWORD,
        }
    return result


@router.patch("/{worker_id}", response_model=WorkerOut)
def update_worker(
    worker_id: str,
    payload: WorkerUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    worker = db.query(Worker).filter(
        Worker.id == worker_id,
        Worker.organisation_id == current_user.organisation_id,
    ).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(worker, key, value)

    db.commit()
    db.refresh(worker)
    return worker


@router.delete("/{worker_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_worker(
    worker_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_staff),
):
    worker = db.query(Worker).filter(
        Worker.id == worker_id,
        Worker.organisation_id == current_user.organisation_id,
    ).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    db.delete(worker)
    db.commit()
